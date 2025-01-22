import pandas as pd 
from openpyxl import load_workbook 
import os  
from xml.etree import ElementTree as ET
from zipfile import ZipFile
import argparse

def extract_images_from_xlsx(input_file, image_output_dir):
    """ 
    Extracts images from an Excel file and saves them to a directory. 
        Args: 
            input_file (str): Path to the Excel file. 
            image_output_dir (str): Directory to save the extracted images. 
        Returns: 
            dict: A dictionary mapping sheet names to lists of image filenames. 
    """ 
    
    if not os.path.exists(image_output_dir): 
        os.makedirs(image_output_dir) 
    sheet_images = {} 

    with ZipFile(input_file, 'r') as archive:
        media_files = [f for f in archive.namelist() if f.startswith("xl/media/")]
        drawing_files = [f for f in archive.namelist() if f.startswith("xl/drawings/")]
        drawing_rels_files = [f for f in archive.namelist() if f.startswith("xl/drawings/_rels/")]
        worksheet_rels = [f for f in archive.namelist() if f.startswith("xl/worksheets/_rels/")]
        workbook_xml = archive.read("xl/workbook.xml")
        workbook_tree = ET.fromstring(workbook_xml)
        sheet_map = {}
        drawing_to_sheets = {}

        for sheet in workbook_tree.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
            sheet_id = sheet.attrib["sheetId"]
            sheet_name = sheet.attrib["name"]
            sheet_map[f"sheet{sheet_id}.xml"] = sheet_name

        print(f"Sheet Mapping: {sheet_map}")

        for rel_path in worksheet_rels:
            internal_sheet_name = rel_path.replace("xl/worksheets/_rels/", "").replace(".xml.rels", ".xml")
            display_sheet_name = sheet_map.get(internal_sheet_name, f"Unknown_Sheet_{internal_sheet_name}")
            rels_xml = archive.read(rel_path)
            rels_tree = ET.fromstring(rels_xml)

            for rel in rels_tree.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                if "drawing" in rel.attrib["Type"]:
                    drawing_file = rel.attrib["Target"].replace("../", "xl/")
                    drawing_to_sheets.setdefault(drawing_file, []).append(display_sheet_name)
        print(f"Drawing to Sheet Mapping: {drawing_to_sheets}")

        for drawing_path in drawing_files:
            #print(drawing_path)
            try:
                drawing_xml = archive.read(drawing_path)
                #print(drawing_xml)
                tree = ET.fromstring(drawing_xml)
                associated_sheets = drawing_to_sheets.get(drawing_path, [f"Unknown_Sheet_{drawing_path}"])
                print(f"Processing Drawing: {drawing_path}, Associated Sheet: {associated_sheets}")

                drawing_rels_path = drawing_path.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
                drawing_rels_xml = archive.read(drawing_rels_path)
                drawing_rels_tree = ET.fromstring(drawing_rels_xml)
                rId_to_media = {
                    rel.attrib["Id"]: rel.attrib["Target"].replace("../", "xl/")
                    for rel in drawing_rels_tree.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship")
                    if "media" in rel.attrib["Target"]
                    }
                for tag in tree.findall(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"):
                    img_rId = tag.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"]
                    media_file = rId_to_media.get(img_rId)
#                    for idx, media_file in enumerate(media_files):
                    if media_file:
                        img_name = os.path.basename(media_file)
                        img_output_path = os.path.join(image_output_dir, img_name)

                        with open(img_output_path, "wb") as img_file:
                            img_file.write(archive.read(media_file))
                            for sheet_name in associated_sheets:
                                sheet_images.setdefault(sheet_name, []).append(img_name)
#                        if f"{sheet_name}_{idx +1}" not in sheet_images:
#                            sheet_images[f"{sheet_name}_{idx +1}"] =[]
#                            sheet_images[f"{sheet_name}_{idx +1}"].append(img_name)
            except Exception as e:
                print(f"Error processing drawing file {drawing_path}: {e}")
    print(f"Final Sheet Images Mapping: {sheet_images}")
    return sheet_images






#        wb = load_workbook(input_file, data_only=True, keep_links=True) 
#       print("Workbook Attributes:")
#       print(dir(wb))

#       for sheet_name in wb.sheetnames:
#           ws = wb[sheet_name]
#           print(f"\nShhet: {sheet_name}")
#           print("Attributes:", dir(ws))
#           if hasattr(ws, "_rels"):
#               print("_rels:", ws._rels)
#           if hasattr(ws, "_images"):
#               print("_rels:", ws._images)
#           if hasattr(ws, "_archive"):
#               print("_rels:", ws._archive)
#       print("\nWorkbook Propertires:")
#       print(vars(wb))

#    for sheet_name in wb.sheetnames: 
#        ws = wb[sheet_name] 
#        images = [] 
        
        # Check for drawing relationships in the sheet
#        if hasattr(ws, "_rels") and ws._rels:
#            for rel in ws._rels:
#                if "drawing" in rel.target:  # Find the drawing relationship
#                    drawing_path = rel.target  # e.g., 'xl/drawings/drawing1.xml'
#                    print(drawing_path)
#                    try:
#                        # Parse the drawing XML
#                        drawing_xml = archive.read(drawing_path)
#                        tree = ET.fromstring(drawing_xml)
#                        # Find image relationships in the drawing XML
#                        for tag in tree.findall(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"):
#                            img_rId = tag.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"]
#                            # Match image relationship ID with the workbook's archive
#                            for rel in ws._rels():
#                                if rel.id == img_rId and "media" in rel.target:
#                                    img_path_in_archive = rel.Target
#                                    img_name = os.path.basename(img_path_in_archive)
#                                    img_output_path = os.path.join(image_output_dir, img_name)
#                                    with open(img_output_path, "wb") as img_file:
#                                        img_file.write(archive.read(img_path_in_archive))
#                                    images.append(img_name)
#                    except KeyError:
#                        print(f"Could not find drawing file: {drawing_path}")
#                    except Exception as e:
#                        print(f"Error processing drawing: {e}")
#        sheet_images[sheet_name] = images
#    return sheet_images 

def convert_xlsx_to_adoc_with_images(input_file, output_file, image_output_dir): 
    """ 
    Converts an Excel file to a Markdown file with data and images. 
        Args: 
            input_file (str): Path to the input .xlsx file. 
            output_file (str): Path to save the output .md file. 
            image_output_dir (str): Directory to save extracted images. 
    """ 
    try: # Extract images 
        images = extract_images_from_xlsx(input_file, image_output_dir) 
        # Load the Excel data 
        excel_data = pd.ExcelFile(input_file) 
        with open(output_file, 'w', encoding='utf-8') as adoc_file: 
            for sheet_name in excel_data.sheet_names: # Write the sheet name 
                adoc_file.write(f"# {sheet_name}\n\n")

                for idx, image in enumerate(images):
                    if sheet_name == image[:-2]:
                        img_path = os.path.join(image_output_dir, f"image{idx +1}.png") 
                        adoc_file.write(f"image::{img_path}[{image}]\n\n") 
                df = excel_data.parse(sheet_name) 
                adoc_file.write("[cols=\"auto\", options=\"header\"]\n|===\n")
                adoc_file.write("| " + " | ".join(df.columns) + "\n")
                for _, row in df.iterrows():
                    adoc_file.write("| " + " | ".join(map(str, row)) + "\n") 
                adoc_file.write("|===\n\n")

        print(f"Successfully converted {input_file} to {output_file} with images in {image_output_dir}") 
    except Exception as e: 
        print(f"Error occurred: {e}") 
        

def main():
    '''
    Main fuction will take two arguments (input_file and output_file) and run a xlsx to asciidoc conversion of the file and put all media in to its own folder "./extracted_images"

    '''
    parser = argparse.ArgumentParser(description="convert xlsx to adoc, including image support")
    parser.add_argument("-i", "--input", required=True, help="xlsx to convert")
    parser.add_argument("-o", "--output", required=True, help="name of file to convert to")

    args = parser.parse_args()
    image_output_dir = "extracted_images/" 
    convert_xlsx_to_adoc_with_images(args.input, args.output, image_output_dir)

if __name__ == "__main__":
    main()
